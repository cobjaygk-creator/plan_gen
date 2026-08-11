"""Step 6 support (still no AI): pull the raw "특별 보상 미리보기" section out
of request.xlsx as a compact grid-preserving text block, ready to hand to
the AI classifier. Everything above this section (title/period/grade
table) was already consumed by step 5's parse_fixed_fields(); everything
from special_reward_start_row down to (but not including) the
"배틀패스 보상 리스트 보러 가기" / "유의사항" boundary belongs here.

Important finding: item captions are NOT always plain cell values. Some
months (confirmed: 202509) put the caption as a floating drawing textbox
anchored near the image instead of a cell value — openpyxl's cell API
sees nothing there. So this module reads text from both sources (cell
values + drawing textbox runs) and merges them by anchor row, the same
way images will need to be position-matched in step 7.
"""
import html
import re
import zipfile
import openpyxl

END_MARKERS_RE = re.compile(r"(배틀패스\s*보상\s*리스트\s*보러\s*가기|^유의사항$)")
COLS = ["B", "C", "D", "E", "F"]

ANCHOR_SPLIT_RE = re.compile(r"(?=<xdr:(?:twoCellAnchor|oneCellAnchor)[ >])")
FROM_RE = re.compile(r"<xdr:from><xdr:col>(\d+)</xdr:col>.*?<xdr:row>(\d+)</xdr:row>", re.S)
FROM_COL_OFF_RE = re.compile(r"<xdr:from><xdr:col>(\d+)</xdr:col>.*?<xdr:colOff>(\d+)</xdr:colOff>", re.S)
TEXT_RUN_RE = re.compile(r"<a:t>(.*?)</a:t>")


def _find_drawing_path(path: str) -> str | None:
    with zipfile.ZipFile(path) as z:
        rels_path = "xl/worksheets/_rels/sheet1.xml.rels"
        if rels_path in z.namelist():
            rels = z.read(rels_path).decode("utf-8")
            m = re.search(r'Target="\.\./(drawings/drawing\d+\.xml)"', rels)
            if m:
                return f"xl/{m.group(1)}"
        candidates = [n for n in z.namelist() if re.match(r"xl/drawings/drawing\d+\.xml$", n)]
        return candidates[0] if candidates else None


def extract_drawing_text_anchors(path: str) -> list[dict]:
    """Returns [{"row": excel_row_1indexed, "col_letter": "B", "col_off": int,
    "text": str}, ...] for every floating textbox (not picture) in the
    sheet's drawing, ignoring shapes that have no text runs at all.

    col_off is the raw <xdr:colOff> EMU offset within col_letter's column —
    kept so callers can tell apart two textboxes that share the same
    col_idx0 (and so the same col_letter) but sit at very different
    horizontal positions. Real case (202508 "배틀패스 의상 교환권"): a
    4-item visual row only spans 2-3 real spreadsheet columns, so e.g.
    "화이트 니트 세트" and "브라운 니트 세트" both anchor at col_idx0=2
    (colOff 539213 vs 2035635 — clearly different cards) — treating them
    as the same column silently dropped one (see extract_special_reward_rows)."""
    drawing_path = _find_drawing_path(path)
    if drawing_path is None:
        return []
    with zipfile.ZipFile(path) as z:
        data = z.read(drawing_path).decode("utf-8", errors="replace")

    anchors = []
    for chunk in ANCHOR_SPLIT_RE.split(data)[1:]:
        if "<xdr:pic>" in chunk:
            continue  # picture anchor, handled separately in step 7
        texts = TEXT_RUN_RE.findall(chunk)
        if not texts:
            continue
        m = FROM_RE.search(chunk)
        if not m:
            continue
        col_idx0, row_idx0 = int(m.group(1)), int(m.group(2))
        if col_idx0 >= len(COLS):
            continue
        off_m = FROM_COL_OFF_RE.search(chunk)
        col_off = int(off_m.group(2)) if off_m else 0
        # XML text runs keep entities literal ("&amp;" for a real "&") —
        # real case: 202508 "빌브라트 수트 &amp; 펀치 mini" rendered with
        # the escape still in it before this unescape was added.
        joined = html.unescape("".join(texts)).strip()
        if joined:
            anchors.append({
                "row": row_idx0 + 1, "col_letter": COLS[col_idx0], "col_off": col_off, "text": joined,
            })
    return anchors


def extract_drawing_picture_rows(path: str) -> set[int]:
    """Row numbers (1-indexed, inclusive of anchor span) that have a
    <xdr:pic> anchored to them. Cheap presence-only check — full position
    matching (which picture belongs to which item) is step 7's job, but
    the classifier needs to know "does this row have an image at all" to
    tell grid apart from text_list (identical as pure text otherwise —
    see 202508, which text-only classified as text_list despite actually
    being an image grid in the rendered output)."""
    drawing_path = _find_drawing_path(path)
    if drawing_path is None:
        return set()
    with zipfile.ZipFile(path) as z:
        data = z.read(drawing_path).decode("utf-8", errors="replace")

    rows_with_pics = set()
    for chunk in ANCHOR_SPLIT_RE.split(data)[1:]:
        if "<xdr:pic>" not in chunk:
            continue
        from_m = FROM_RE.search(chunk)
        to_m = re.search(r"<xdr:to>.*?<xdr:row>(\d+)</xdr:row>", chunk, re.S)
        if not from_m:
            continue
        row_start = int(from_m.group(2)) + 1
        row_end = int(to_m.group(1)) + 1 if to_m else row_start
        for r in range(row_start, row_end + 1):
            rows_with_pics.add(r)
    return rows_with_pics


def extract_special_reward_rows(path: str, start_row: int) -> list[dict]:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    end_row = ws.max_row + 1
    for r in range(start_row + 1, ws.max_row + 1):
        for c in range(2, 7):
            v = ws.cell(row=r, column=c).value
            if v and END_MARKERS_RE.search(str(v).strip()):
                end_row = r
                break
        if end_row == r:
            break

    pic_rows = extract_drawing_picture_rows(path)

    row_map: dict[int, dict] = {}

    for r in range(start_row, end_row):
        cells = {}
        for col_idx, col_letter in zip(range(2, 7), COLS):
            v = ws.cell(row=r, column=col_idx).value
            if v not in (None, ""):
                cells[col_letter] = str(v).strip()
        # a row can be worth keeping even with zero cell text, if a picture
        # is anchored there and nothing else describes it — e.g. a pure
        # icon-choice section (real case: 202602 "고양이 모자 선택권", whose
        # 9 icons are one combined image anchored to rows with no text in
        # any cell at all). Dropping these silently meant the classifier
        # never even knew the section had content beyond its title/footnote.
        if cells or r in pic_rows:
            row_map[r] = {"row": r, "cells": cells}

    text_anchors_by_row: dict[int, list[dict]] = {}
    for anchor in extract_drawing_text_anchors(path):
        r = anchor["row"]
        if not (start_row <= r < end_row):
            continue
        text_anchors_by_row.setdefault(r, []).append(anchor)

    for r, row_anchors in text_anchors_by_row.items():
        entry = row_map.setdefault(r, {"row": r, "cells": {}})
        # process left-to-right by true position (col_idx0, then colOff) so
        # that when two textboxes round to the same col_letter (see real
        # 202508 case in extract_drawing_text_anchors' docstring), the
        # later one gets bumped to the nearest free letter instead of
        # silently overwriting/being dropped — every caption survives.
        row_anchors.sort(key=lambda a: (COLS.index(a["col_letter"]), a["col_off"]))
        for anchor in row_anchors:
            letter = anchor["col_letter"]
            if letter in entry["cells"] and entry["cells"][letter] != anchor["text"]:
                start_idx = COLS.index(letter)
                free = next(
                    (c for c in COLS[start_idx + 1:] if c not in entry["cells"]),
                    next((c for c in COLS if c not in entry["cells"]), None),
                )
                if free is None:
                    continue  # row already has one caption in every COLS slot — drop, same as before
                letter = free
            # don't clobber a real cell value with a textbox guess at the same slot
            entry["cells"].setdefault(letter, anchor["text"])

    for r, entry in row_map.items():
        entry["has_image"] = r in pic_rows

    return [row_map[r] for r in sorted(row_map)]


def to_compact_text(rows: list[dict]) -> str:
    """Grid-preserving compact text — cheap on tokens, keeps enough
    structure (which columns are populated together) for the model to
    infer N-column grid vs single-column list vs paired layout. Appends
    an explicit image-presence marker per row since that's otherwise
    invisible in plain text and is the deciding signal between grid and
    text_list (see extract_drawing_picture_rows)."""
    lines = []
    for row in rows:
        cells = row["cells"]
        cols = "".join(cells.keys())
        values = " | ".join(cells.values())
        marker = " [이미지있음]" if row.get("has_image") else ""
        if values:
            lines.append(f"{cols}{row['row']}: {values}{marker}")
        else:
            # no cell text at all, but kept because a picture anchors here
            # (see extract_special_reward_rows) — render as a bare marker
            # line so the classifier can still see "there's content here"
            lines.append(f"{row['row']}:{marker}")
    return "\n".join(lines)


if __name__ == "__main__":
    import sys
    from tools.parse_fixed_fields import parse_fixed_fields

    path = sys.argv[1] if len(sys.argv) > 1 else "samples/202605_request.xlsx"
    fixed = parse_fixed_fields(path)
    rows = extract_special_reward_rows(path, fixed["special_reward_start_row"])
    text = to_compact_text(rows)
    with open("templates/_extract_debug.txt", "w", encoding="utf-8") as f:
        f.write(text + f"\n\n--- {len(rows)} rows, {sum(len(r['cells']) for r in rows)} cells ---")
