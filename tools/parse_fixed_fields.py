"""Step 5: parse request.xlsx fixed fields (title/period/grade table/notices)
with regex — no AI. These are the same every month in structure, but their
row numbers shift depending on how many "특별 보상" rows sit above them, so
every field is found by searching for its label/pattern rather than by a
fixed cell address.

The "특별 보상 미리보기" section itself (item names + images) is NOT parsed
here — that's step 6's AI classification job, this module explicitly stops
at its header and leaves everything below it to that stage.
"""
import re
import openpyxl

PERIOD_RE = re.compile(r".*점검.*~.*(AM|PM|오전|오후).*\d{1,2}:\d{2}.*")
GRADE_HEADER_RE = re.compile(r"배틀패스\s*분류")
SPECIAL_REWARD_HEADER_RE = re.compile(r"특별\s*보상\s*미리보기")
NOTICE_HEADER_RE = re.compile(r"^유의사항$")


def _iter_cells(ws):
    for row in ws.iter_rows():
        for cell in row:
            if cell.value not in (None, ""):
                yield cell


def parse_fixed_fields(path: str) -> dict:
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb["Sheet1"]

    period_row = None
    period_text = None
    grade_row = None
    special_reward_row = None
    notice_row = None

    for cell in _iter_cells(ws):
        text = str(cell.value).strip()
        if period_row is None and PERIOD_RE.match(text):
            period_row, period_text = cell.row, text
        if grade_row is None and GRADE_HEADER_RE.search(text):
            grade_row = cell.row
        if special_reward_row is None and SPECIAL_REWARD_HEADER_RE.search(text):
            special_reward_row = cell.row
        if notice_row is None and NOTICE_HEADER_RE.match(text):
            notice_row = cell.row

    if period_row is None:
        raise ValueError(f"{path}: could not find period text (label search failed)")
    if grade_row is None:
        raise ValueError(f"{path}: could not find grade table header '배틀패스 분류'")

    # Title: every non-empty column-B cell strictly above the period row.
    title_lines = []
    for r in range(1, period_row):
        v = ws.cell(row=r, column=2).value
        if v not in (None, ""):
            title_lines.append(str(v).strip())

    # Grade table: label row, then tier-name row, then tier-price row.
    tier_row = grade_row + 1
    price_row = grade_row + 2
    tiers = [ws.cell(row=tier_row, column=c).value for c in (2, 3, 4)]
    prices = [ws.cell(row=price_row, column=c).value for c in (2, 3, 4)]
    if any(t in (None, "") for t in tiers) or any(p in (None, "") for p in prices):
        raise ValueError(f"{path}: grade table rows incomplete near row {grade_row}")
    grade_table = {
        "tiers": [str(t).strip() for t in tiers],
        "prices": [str(p).strip() for p in prices],
    }

    # Notices: every non-empty column-B cell after the "유의사항" header.
    notices = []
    if notice_row is not None:
        for r in range(notice_row + 1, ws.max_row + 1):
            v = ws.cell(row=r, column=2).value
            if v not in (None, ""):
                notices.append(str(v).strip())

    return {
        "title_lines": title_lines,
        "period": period_text,
        "grade_table": grade_table,
        "notices": notices,
        "special_reward_start_row": special_reward_row,  # handoff point for step 6
    }


if __name__ == "__main__":
    import sys
    import json
    path = sys.argv[1] if len(sys.argv) > 1 else "samples/202605_request.xlsx"
    result = parse_fixed_fields(path)
    print(json.dumps(result, ensure_ascii=False, indent=2))
